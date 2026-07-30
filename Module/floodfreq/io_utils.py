"""
Input/output helpers: read an annual-maximum series from CSV or Excel, and
write a FloodFrequencyAnalysis's results back out to a formatted .xlsx
report (so results stay easy to share with people who work in Excel).

Also provides project-layout helpers for the standard folder structure:

    <project_root>/
        Data/<CaseName>.csv     - input data, one file per case
        Module/                 - this package + scripts (Module/floodfreq/...)
        Output/<CaseName>/      - CSV outputs for that case
        Plot/<CaseName>/        - PNG plots for that case

`resolve_case()` derives all of these paths from a case name, given the path
of the calling script (pass `__file__` from a script living in Module/).
"""
from __future__ import annotations
from dataclasses import dataclass
from pathlib import Path
import tomllib
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


@dataclass
class CasePaths:
    case_name: str
    project_root: Path
    data_csv: Path
    output_dir: Path
    plot_dir: Path


@dataclass
class RegionPaths:
    """Standard path layout for one regional (pooled) analysis group:

        <project_root>/
            Data/Regional/<RegionName>/*.csv   - one file per station
            Output/Regional/<RegionName>/      - CSV outputs for this region
            Plot/Regional/<RegionName>/         - PNG plots for this region
    """
    region_name: str
    project_root: Path
    data_dir: Path
    output_dir: Path
    plot_dir: Path


@dataclass
class ClimatePaths:
    """Standard path layout for one climate-informed (CIFAM) adjustment case:

        <project_root>/
            Data/<CaseName>/<CaseName>.csv                      - baseline series
                                                                  (reused from the
                                                                  single-station tool)
            Data/<CaseName>/climate_adjustment/<scenario>.csv   - one climate-input
                                                                  file per scenario
            Output/<CaseName>/Climate_Adjustment/<scenario>/    - CSV outputs
            Plot/<CaseName>/Climate_Adjustment/<scenario>/      - PNG plots

    Everything is CASE-FIRST and, within a case, one <scenario> per climate
    input file, so several scenarios (e.g. rcp45, rcp85, wet, dry) coexist
    without overwriting each other's outputs. The baseline series is the same
    file the single-station tool uses (Data/<CaseName>/<CaseName>.csv), so a
    case already analysed with run_analysis.py needs no data duplication.
    Regional analysis stays analysis-type-first (Output/Regional/<RegionName>/)
    on purpose: a region isn't a single case, whereas a climate adjustment is
    an extension of one.
    """
    case_name: str
    scenario: str
    project_root: Path
    baseline_csv: Path       # Data/<CaseName>/<CaseName>.csv                    (baseline)
    climate_csv: Path        # Data/<CaseName>/climate_adjustment/<scenario>.csv (the 4 params)
    output_dir: Path         # Output/<CaseName>/Climate_Adjustment/<scenario>/
    plot_dir: Path           # Plot/<CaseName>/Climate_Adjustment/<scenario>/


def project_root_from(module_file) -> Path:
    """
    Given __file__ of a script living directly under <project_root>/Module/,
    return <project_root>. Walks up until a directory containing both a
    'Data' and a 'Module' subdirectory is found, so it also works for
    scripts nested one level deeper inside Module/.
    """
    p = Path(module_file).resolve().parent
    for candidate in [p] + list(p.parents):
        if (candidate / "Data").is_dir() and (candidate / "Module").is_dir():
            return candidate
    # Fallback: assume the script sits directly in Module/
    return p.parent


def resolve_case(case_name: str, module_file) -> CasePaths:
    """
    Resolve the standard Data/Module/Output/Plot paths for one case, and
    create Output/<CaseName>/ and Plot/<CaseName>/ if they don't exist yet.

    Case-first Data layout (v0.8.0): each case owns a folder
    Data/<CaseName>/ containing its baseline series Data/<CaseName>/<CaseName>.csv
    (named after the folder, by convention) and, optionally, a per-case
    settings file Data/<CaseName>/<CaseName>.toml and a
    Data/<CaseName>/climate_adjustment/ subfolder (see resolve_climate_case).
    This mirrors the case-first Output/<CaseName>/ and Plot/<CaseName>/ trees.
    """
    root = project_root_from(module_file)
    data_csv = root / "Data" / case_name / f"{case_name}.csv"
    output_dir = root / "Output" / case_name
    plot_dir = root / "Plot" / case_name
    output_dir.mkdir(parents=True, exist_ok=True)
    plot_dir.mkdir(parents=True, exist_ok=True)
    return CasePaths(case_name, root, data_csv, output_dir, plot_dir)


def resolve_region(region_name: str, module_file) -> RegionPaths:
    """
    Resolve the standard Data/Regional/Output/Regional/Plot/Regional paths
    for one regional (pooled) analysis group, and create
    Output/Regional/<RegionName>/ and Plot/Regional/<RegionName>/ if they
    don't exist yet.
    """
    root = project_root_from(module_file)
    data_dir = root / "Data" / "Regional" / region_name
    output_dir = root / "Output" / "Regional" / region_name
    plot_dir = root / "Plot" / "Regional" / region_name
    output_dir.mkdir(parents=True, exist_ok=True)
    plot_dir.mkdir(parents=True, exist_ok=True)
    return RegionPaths(region_name, root, data_dir, output_dir, plot_dir)


def resolve_climate_case(case_name: str, scenario: str, module_file) -> ClimatePaths:
    """
    Resolve the standard paths for one climate-informed (CIFAM) adjustment
    scenario of one case, and create the per-scenario output/plot dirs.

    Case-first, scenario-nested (v0.8.0):
        baseline   Data/<CaseName>/<CaseName>.csv           (reused as-is)
        inputs     Data/<CaseName>/climate_adjustment/<scenario>.csv
        outputs    Output/<CaseName>/Climate_Adjustment/<scenario>/
        plots      Plot/<CaseName>/Climate_Adjustment/<scenario>/

    Several scenarios per case (rcp45, rcp85, ...) coexist without clobbering
    each other's outputs. The baseline is the single-station input, so no
    duplication. Regional analysis stays analysis-type-first
    (Output/Regional/<RegionName>/) on purpose: a region isn't a single case,
    whereas a climate adjustment is an extension of one.
    """
    root = project_root_from(module_file)
    baseline_csv = root / "Data" / case_name / f"{case_name}.csv"
    climate_csv = root / "Data" / case_name / "climate_adjustment" / f"{scenario}.csv"
    output_dir = root / "Output" / case_name / "Climate_Adjustment" / scenario
    plot_dir = root / "Plot" / case_name / "Climate_Adjustment" / scenario
    output_dir.mkdir(parents=True, exist_ok=True)
    plot_dir.mkdir(parents=True, exist_ok=True)
    return ClimatePaths(case_name, scenario, root, baseline_csv, climate_csv,
                        output_dir, plot_dir)


# Keys expected in a Data/<CaseName>/climate_adjustment/<scenario>.csv inputs file.
_CLIMATE_REQUIRED = ("delta1", "delta2", "tau1", "tau2")
_CLIMATE_OPTIONAL_DEFAULTS = {
    "confidence_level": 95.0,
    "distribution": "gumbel",
    "return_periods": (2, 5, 10, 20, 50, 100, 200, 500, 1000, 2000, 5000, 10000, 100000),
    "pmf": None,
}


def load_climate_inputs(path) -> dict:
    """
    Read the four CIFAM climate numbers (+ options) from a long-format CSV:

        parameter,value,units,description
        delta1,0.30,fraction,...
        delta2,0.18,fraction,...
        tau1,0.10,fraction,...
        tau2,0.18,fraction,...
        confidence_level,95,percent,...
        return_periods,"2,5,...,100000",years,...
        distribution,gumbel,name,...
        pmf,,m3/s,...

    Only the 'parameter' and 'value' columns are read; 'units'/'description'
    are documentation for whoever edits the file and are ignored here. The
    long "one parameter per row" format (rather than a single wide header
    row) is used so the description column can travel with each value --
    consistent with how the descriptor-catalog templates are laid out, and
    friendlier for a non-developer duplicating a template per scenario.

    delta1/delta2/tau1/tau2 are REQUIRED and are FRACTIONS (0.30 == +30%);
    a common mistake is entering 30 (meaning percent) -- values with
    magnitude > 1.5 raise, to catch that early rather than producing absurd
    floods silently. Returns a dict with keys delta1, delta2, tau1, tau2,
    confidence_level, distribution, return_periods (tuple), pmf (float|None).
    """
    path = str(path)
    if not Path(path).exists():
        raise FileNotFoundError(
            f"Climate inputs file not found: {path}\n"
            f"Copy a climate-adjustment template from Data/Templates/ into "
            f"Data/<CaseName>/climate_adjustment/<scenario>.csv and edit the "
            f"four delta/tau values."
        )
    df = pd.read_csv(path)
    if "parameter" not in df.columns or "value" not in df.columns:
        raise ValueError(
            f"{path} must have 'parameter' and 'value' columns "
            f"(see the climate-adjustment template in Data/Templates/)."
        )
    raw = {str(k).strip(): v for k, v in zip(df["parameter"], df["value"])}

    out = {}
    for key in _CLIMATE_REQUIRED:
        if key not in raw or pd.isna(raw[key]):
            raise ValueError(f"{path}: required parameter '{key}' is missing or blank.")
        val = float(raw[key])
        if abs(val) > 1.5:
            raise ValueError(
                f"{path}: '{key}' = {val} looks like a percent, not a fraction. "
                f"Enter fractions (e.g. 0.30 for +30%), not 30."
            )
        out[key] = val

    # optional: confidence_level
    cl = raw.get("confidence_level", None)
    out["confidence_level"] = (float(cl) if cl is not None and not pd.isna(cl)
                               else _CLIMATE_OPTIONAL_DEFAULTS["confidence_level"])

    # optional: distribution
    dist = raw.get("distribution", None)
    out["distribution"] = (str(dist).strip() if dist is not None and not pd.isna(dist)
                           else _CLIMATE_OPTIONAL_DEFAULTS["distribution"])

    # optional: return_periods (comma-separated string, or a single number)
    rp = raw.get("return_periods", None)
    if rp is None or (isinstance(rp, float) and pd.isna(rp)):
        out["return_periods"] = tuple(_CLIMATE_OPTIONAL_DEFAULTS["return_periods"])
    else:
        out["return_periods"] = tuple(
            float(x) for x in str(rp).replace(";", ",").split(",") if str(x).strip()
        )

    # optional: pmf reference value
    pmf = raw.get("pmf", None)
    out["pmf"] = (float(pmf) if pmf is not None and not pd.isna(pmf)
                  and str(pmf).strip() != "" else None)

    return out


def load_region_stations(paths: RegionPaths, value_col=None, year_col=None):
    """
    Read every station CSV in Data/Regional/<RegionName>/ into two dicts
    keyed by station name (the file's stem, e.g.
    Data/Regional/<RegionName>/P1009.csv -> "P1009"):

        station_data:  {station_name: values_array}
        station_years: {station_name: years_array or None}

    `station_years[name]` is None for any station CSV without a detectable
    year column -- perfectly usable (data_quality/Mann-Kendall fall back to
    record order as a stand-in for time, same as the single-station tool
    does), but it does mean the "missing year" validation warning and
    Sen's-slope-per-calendar-year magnitude aren't available for that
    station. `.toml` files in the same folder (if any are ever added) are
    ignored here -- unlike single-station cases, regional stations don't
    currently support per-station settings files.
    """
    if not paths.data_dir.is_dir():
        raise FileNotFoundError(
            f"Region data folder not found: {paths.data_dir}\n"
            f"Create it and add one CSV per station, e.g. "
            f"{paths.data_dir}/StationA.csv")
    csvs = sorted(paths.data_dir.glob("*.csv"))
    if not csvs:
        raise ValueError(
            f"No station CSVs found in {paths.data_dir}. Add at least 4 "
            f"(one per station) to run a regional analysis.")
    station_data = {}
    station_years = {}
    for f in csvs:
        values, years = read_series(f, value_col=value_col, year_col=year_col)
        station_data[f.stem] = values
        station_years[f.stem] = years
    return station_data, station_years


def load_case_config(case_name: str, project_root: Path) -> dict:
    """
    Read optional per-case settings from Data/<CaseName>/<CaseName>.toml, if
    it exists. Returns {} if the file doesn't exist (config is entirely
    optional; a case with no .toml behaves exactly as it always has).

    Recognized keys mirror run_analysis.py's CLI flag names (with dashes
    replaced by underscores, as TOML keys can't contain dashes as bare
    identifiers): value_col, year_col, variable_name, units, short_name,
    plotting_position, descriptive_plotting_position, n_boot,
    confidence_level, regional_skew, regional_skew_mse, no_plots,
    xlsx_report, pdf_report. Unrecognized keys are ignored (not an error),
    so a config file can be extended later without breaking older runs of
    the tool against it.
    """
    config_path = project_root / "Data" / case_name / f"{case_name}.toml"
    if not config_path.exists():
        return {}
    try:
        with open(config_path, "rb") as f:
            return tomllib.load(f)
    except tomllib.TOMLDecodeError as e:
        raise ValueError(f"Could not parse {config_path}: {e}") from e


def read_series(path, value_col=None, year_col=None, sheet_name=0):
    """
    Read an annual-maximum series from a CSV or Excel file.

    If value_col/year_col are not given, the function guesses: the first
    all-numeric column with more than one distinct value is treated as the
    flood series, and a column that looks like a 4-digit year is used as
    the year index if present.
    """
    path_str = str(path)
    if not Path(path_str).exists():
        raise FileNotFoundError(f"Input file not found: {path_str}")

    if path_str.lower().endswith((".csv",)):
        df = pd.read_csv(path_str)
    else:
        df = pd.read_excel(path_str, sheet_name=sheet_name)

    if df.empty or df.shape[1] == 0:
        raise ValueError(f"{path_str} was read successfully but contains no data "
                          f"(no rows or no columns). Check the file isn't empty or "
                          f"malformed.")

    if value_col is not None and value_col not in df.columns:
        raise ValueError(f"value_col='{value_col}' not found in {path_str}. "
                          f"Available columns: {list(df.columns)}")
    if year_col is not None and year_col not in df.columns:
        raise ValueError(f"year_col='{year_col}' not found in {path_str}. "
                          f"Available columns: {list(df.columns)}")

    if value_col is None:
        numeric_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
        if not numeric_cols:
            raise ValueError(f"No numeric column found in {path_str}; pass value_col "
                              f"explicitly. Available columns: {list(df.columns)}")
        # prefer a column that is NOT a plausible year column
        candidates = [c for c in numeric_cols
                      if not (df[c].between(1800, 2100).mean() > 0.9)]
        value_col = candidates[0] if candidates else numeric_cols[-1]

    if year_col is None:
        for c in df.columns:
            if pd.api.types.is_numeric_dtype(df[c]) and df[c].between(1800, 2100).mean() > 0.9:
                year_col = c
                break

    sub = df[[year_col, value_col]] if year_col else df[[value_col]]
    n_before = len(sub)
    sub = sub.dropna()
    if sub.empty:
        raise ValueError(f"After dropping missing values, no data remains from column "
                          f"'{value_col}' in {path_str} (started with {n_before} row(s)). "
                          f"Check that column actually contains the flood series.")

    values = sub[value_col].to_numpy(dtype=float)
    years = sub[year_col].to_numpy() if year_col else None
    return values, years


def write_report(path, ffa, return_periods=(2, 5, 10, 20, 25, 50, 100, 200, 500, 1000)):
    """Write descriptive stats, goodness-of-fit ranking, and a quantile
    table to a formatted .xlsx workbook."""
    wb = Workbook()

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    body_font = Font(name="Arial")
    title_font = Font(name="Arial", bold=True, size=14)

    def _write_table(ws, df, start_row=1, start_col=1):
        for j, col in enumerate(df.columns):
            cell = ws.cell(row=start_row, column=start_col + j, value=str(col))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for i, (_, row) in enumerate(df.iterrows()):
            for j, col in enumerate(df.columns):
                val = row[col]
                if isinstance(val, (np.floating, float)):
                    val = round(float(val), 3)
                elif isinstance(val, (np.integer,)):
                    val = int(val)
                cell = ws.cell(row=start_row + 1 + i, column=start_col + j, value=val)
                cell.font = body_font
        for j, col in enumerate(df.columns):
            ws.column_dimensions[get_column_letter(start_col + j)].width = max(14, len(str(col)) + 2)

    # -- Summary sheet --
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"{ffa.short_name} Frequency Analysis — {ffa.station_id}"
    ws["A1"].font = title_font
    stats = ffa.descriptive_stats()
    rows = [(k, v) for k, v in stats.items()]
    stats_df = pd.DataFrame(rows, columns=["Statistic", "Value"])
    _write_table(ws, stats_df, start_row=3)

    # -- Goodness of fit --
    ws2 = wb.create_sheet("Goodness of fit")
    gof = ffa.goodness_of_fit_table()
    _write_table(ws2, gof, start_row=1)

    # -- Quantile table --
    ws3 = wb.create_sheet("Quantiles")
    q = ffa.quantile_table(return_periods=return_periods)
    _write_table(ws3, q, start_row=1)

    # -- Raw data --
    ws4 = wb.create_sheet("Data")
    data_df = pd.DataFrame({"year": ffa.years, "Q": ffa.data}) if ffa.years is not None \
        else pd.DataFrame({"Q": ffa.data})
    _write_table(ws4, data_df, start_row=1)

    wb.save(path)
    return path
